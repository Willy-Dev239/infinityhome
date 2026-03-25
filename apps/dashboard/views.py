import functools
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.auth.models import User
from apps.demandes.models import Demande, Notification, STATUTS
from apps.techniciens.models import Technicien, SPECIALITES, DISPONIBILITE
from apps.accounts.models import Profil



from django.contrib.auth.decorators import login_required, user_passes_test

from django.core.paginator import Paginator
from django.db.models import Q, Sum
from apps.demandes.models import Paiement, InstructionPaiement

is_admin = lambda u: u.is_staff or u.is_superuser

TYPE_TO_SPEC = {
    'electricite': 'electricien',
    'plomberie':   'plombier',
    'construction':'ingenieur',
    'soudure':     'soudeur',
    'peinture':    'peintre',
    'menuiserie':  'menuisier',
    'surveillance':'surveillance',
}

def admin_required(view_func):
    @functools.wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return redirect('dashboard:login')
        return view_func(request, *args, **kwargs)
    return wrapper

def admin_login(request):
    if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
        return redirect('dashboard:home')
    error = None
    if request.method == 'POST':
        user = authenticate(request, username=request.POST.get('username',''), password=request.POST.get('password',''))
        if user and (user.is_staff or user.is_superuser):
            login(request, user)
            return redirect('dashboard:home')
        error = 'Identifiants incorrects ou acces refuse.'
    return render(request, 'dashboard/login.html', {'error': error})

def admin_logout(request):
    logout(request)
    return redirect('dashboard:login')

@admin_required
def home(request):
    stats = {
        'total_demandes':   Demande.objects.count(),
        'en_attente':       Demande.objects.filter(statut='en_attente').count(),
        'en_cours':         Demande.objects.filter(statut__in=['assignee','en_cours']).count(),
        'terminees':        Demande.objects.filter(statut='terminee').count(),
        'total_techniciens':Technicien.objects.filter(actif=True).count(),
        'disponibles':      Technicien.objects.filter(disponibilite='disponible', actif=True).count(),
        'total_clients':    User.objects.filter(is_staff=False).count(),
    }
    return render(request, 'dashboard/home.html', {
        'stats': stats,
        'demandes_recentes': Demande.objects.order_by('-date_creation')[:8],
        'techniciens_dispo': Technicien.objects.filter(disponibilite='disponible', actif=True)[:6],
    })

@admin_required
def demandes(request):
    statut = request.GET.get('statut', '')
    qs = Demande.objects.select_related('client','technicien').order_by('-date_creation')
    if statut:
        qs = qs.filter(statut=statut)
    return render(request, 'dashboard/demandes.html', {'demandes': qs, 'statut_actif': statut})

@admin_required
def demande_detail(request, pk):
    demande = get_object_or_404(Demande, pk=pk)
    specialite = TYPE_TO_SPEC.get(demande.type_travaux, demande.type_travaux)
    techniciens_dispo = Technicien.objects.filter(disponibilite='disponible', actif=True, specialite=specialite)
    tous_techniciens  = Technicien.objects.filter(actif=True, specialite=specialite)
    return render(request, 'dashboard/demande_detail.html', {
        'demande': demande,
        'techniciens_dispo': techniciens_dispo,
        'tous_techniciens': tous_techniciens,
        'statuts': STATUTS,
    })

@admin_required
def assigner_technicien(request, pk):
    demande = get_object_or_404(Demande, pk=pk)
    if request.method == 'POST':
        tech_id = request.POST.get('technicien_id')
        date_intervention = request.POST.get('date_intervention')
        if tech_id:
            tech = get_object_or_404(Technicien, pk=tech_id)
            demande.technicien = tech
            demande.statut = 'assignee'
            if date_intervention:
                demande.date_intervention = date_intervention
            demande.save()
            Notification.objects.create(
                utilisateur=demande.client,
                demande=demande,
                titre='Technicien assigne a votre demande',
                message=f'Bonjour {demande.client.first_name}, {tech.prenom} {tech.nom} ({tech.get_specialite_display()}) a ete assigne. Tel: {tech.telephone}',
            )
            messages.success(request, f'Technicien {tech.nom_complet} assigne !')
        else:
            messages.error(request, 'Veuillez choisir un technicien.')
    return redirect('dashboard:demande_detail', pk=pk)

@admin_required
def update_statut(request, pk):
    demande = get_object_or_404(Demande, pk=pk)
    if request.method == 'POST':
        new_statut = request.POST.get('statut')
        if new_statut:
            demande.statut = new_statut
            demande.save()
            Notification.objects.create(
                utilisateur=demande.client,
                demande=demande,
                titre='Statut mis a jour',
                message=f'Votre demande "{demande.titre}" est maintenant : {demande.get_statut_display()}.',
            )
            messages.success(request, 'Statut mis a jour.')
    return redirect('dashboard:demande_detail', pk=pk)

@admin_required
def techniciens(request):
    return render(request, 'dashboard/techniciens.html', {
        'techniciens': Technicien.objects.all().order_by('specialite','nom'),
    })

@admin_required
def technicien_add(request):
    if request.method == 'POST':
        try:
            t = Technicien(
                nom=request.POST['nom'], prenom=request.POST['prenom'],
                specialite=request.POST['specialite'], telephone=request.POST['telephone'],
                email=request.POST.get('email',''), localite=request.POST['localite'],
                disponibilite=request.POST.get('disponibilite','disponible'),
                experience=request.POST.get('experience', 0),
            )
            if 'photo' in request.FILES:
                t.photo = request.FILES['photo']
            t.save()
            messages.success(request, f'Technicien {t.nom_complet} ajoute !')
            return redirect('dashboard:techniciens')
        except Exception as e:
            messages.error(request, f'Erreur: {e}')
    return render(request, 'dashboard/technicien_form.html', {
        'specialites': SPECIALITES, 'disponibilites': DISPONIBILITE
    })

@admin_required
def technicien_edit(request, pk):
    tech = get_object_or_404(Technicien, pk=pk)
    if request.method == 'POST':
        tech.nom=request.POST['nom']; tech.prenom=request.POST['prenom']
        tech.specialite=request.POST['specialite']; tech.telephone=request.POST['telephone']
        tech.email=request.POST.get('email',''); tech.localite=request.POST['localite']
        tech.disponibilite=request.POST.get('disponibilite','disponible')
        tech.experience=request.POST.get('experience',0); tech.actif='actif' in request.POST
        if 'photo' in request.FILES:
            tech.photo = request.FILES['photo']
        tech.save()
        messages.success(request, f'Technicien {tech.nom_complet} mis a jour !')
        return redirect('dashboard:techniciens')
    return render(request, 'dashboard/technicien_form.html', {
        'tech': tech, 'specialites': SPECIALITES, 'disponibilites': DISPONIBILITE
    })

@admin_required
def clients(request):
    return render(request, 'dashboard/clients.html', {
        'clients': User.objects.filter(is_staff=False).order_by('-date_joined')
    })
    
    



# ─────────────────────────────────────────────
#  Liste des paiements
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def paiements_list(request):
    qs = Paiement.objects.select_related('client', 'commande').order_by('-created_at')

    filtre_statut = request.GET.get('statut', '')
    filtre_mode   = request.GET.get('mode', '')
    recherche     = request.GET.get('q', '')

    if filtre_statut:
        qs = qs.filter(statut=filtre_statut)
    if filtre_mode:
        qs = qs.filter(mode_paiement=filtre_mode)
    if recherche:
        qs = qs.filter(
            Q(client__username__icontains=recherche) |
            Q(client__first_name__icontains=recherche) |
            Q(client__last_name__icontains=recherche) |
            Q(client__email__icontains=recherche)
        )

    # Stats
    tous = Paiement.objects.all()
    stats = {
        'en_attente':   tous.filter(statut='en_attente').count(),
        'valides':      tous.filter(statut='valide').count(),
        'rejetes':      tous.filter(statut='rejete').count(),
        'total_valide': tous.filter(statut='valide').aggregate(t=Sum('montant'))['t'] or 0,
    }

    paginator = Paginator(qs, 15)
    page      = request.GET.get('page', 1)
    paiements = paginator.get_page(page)

    return render(request, 'dashboard/paiement_list.html', {
        'paiements':      paiements,
        'stats':          stats,
        'filtre_statut':  filtre_statut,
        'filtre_mode':    filtre_mode,
        'recherche':      recherche,
    })


# ─────────────────────────────────────────────
#  Détail d'un paiement
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def paiement_detail(request, pk):
    paiement = get_object_or_404(Paiement, pk=pk)
    instruction = InstructionPaiement.objects.filter(
        mode_paiement=paiement.mode_paiement, actif=True
    ).first()
    return render(request, 'dashboard/paiement_detail.html', {
        'paiement':    paiement,
        'instruction': instruction,
    })


# ─────────────────────────────────────────────
#  Valider un paiement
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def paiement_valider(request, pk):
    paiement = get_object_or_404(Paiement, pk=pk, statut='en_attente')
    paiement.valider(request.user)
    messages.success(request, f'Paiement #{pk} validé avec succès.')
    return redirect('dashboard:paiements')


# ─────────────────────────────────────────────
#  Rejeter un paiement
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def paiement_rejeter(request, pk):
    if request.method == 'POST':
        paiement  = get_object_or_404(Paiement, pk=pk, statut='en_attente')
        note      = request.POST.get('note_admin', '').strip()
        paiement.rejeter(request.user, note=note or None)
        messages.success(request, f'Paiement #{pk} rejeté.')
    return redirect('dashboard:paiements')


# ─────────────────────────────────────────────
#  Configurer les instructions
# ─────────────────────────────────────────────
MODES = [
    ('carte',     'Carte Bancaire',  'credit-card'),
    ('livraison', 'À la Livraison',  'home'),
    ('mobile',    'Mobile Money',    'mobile-alt'),
]

@login_required
@user_passes_test(is_admin)
def instructions_paiement(request):
    instructions = InstructionPaiement.objects.all()

    # Préparer etapes_texte (une étape par ligne) pour chaque instruction
    for inst in instructions:
        inst.etapes_texte = '\n'.join(inst.etapes) if inst.etapes else ''

    return render(request, 'dashboard/instructionpaiement.html', {
        'modes':        MODES,
        'instructions': instructions,
    })


@login_required
@user_passes_test(is_admin)
def instructions_sauvegarder(request, mode):
    if request.method != 'POST':
        return redirect('dashboard:instructions_paiement')

    titre       = request.POST.get('titre', '').strip()
    description = request.POST.get('description', '').strip()
    etapes_raw  = request.POST.get('etapes', '')
    actif       = request.POST.get('actif', '1') == '1'

    # Convertir les lignes en liste
    etapes = [e.strip() for e in etapes_raw.splitlines() if e.strip()]

    InstructionPaiement.objects.update_or_create(
        mode_paiement=mode,
        defaults={
            'titre':       titre,
            'description': description,
            'etapes':      etapes,
            'actif':       actif,
        }
    )
    messages.success(request, f'Instructions "{dict((m[:2] for m in MODES)).get(mode, mode)}" sauvegardées.')
    return redirect('dashboard:instructions_paiement')







from apps.demandes.models import Demande, Paiement
from django.db.models import Count, Sum

stats = {
     # tes stats existantes
    'en_cours':              Demande.objects.filter(statut='en_cours').count(),
    'annulees':              Demande.objects.filter(statut='annulee').count(),
    'paiements_en_attente':  Paiement.objects.filter(statut='en_attente').count(),
}

paiements_recents = Paiement.objects.filter(
    statut='en_attente'
).select_related('client').order_by('-created_at')[:4]


from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Avg, Sum, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncMonth, TruncWeek
from django.utils import timezone
from datetime import timedelta, date
import json

from apps.demandes.models import Demande, Paiement

is_admin = lambda u: u.is_staff or u.is_superuser


@login_required
@user_passes_test(is_admin)
def statistiques(request):
    today = timezone.now().date()
    il_y_a_12_mois = today - timedelta(days=365)
    il_y_a_6_mois  = today - timedelta(days=180)

    # ── 1. Demandes par mois (12 derniers mois) ──
    demandes_par_mois_qs = (
        Demande.objects
        .filter(date_creation__date__gte=il_y_a_12_mois)
        .annotate(mois=TruncMonth('date_creation'))
        .values('mois')
        .annotate(total=Count('id'))
        .order_by('mois')
    )
    mois_labels = []
    mois_data   = []
    mois_fr = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
    for row in demandes_par_mois_qs:
        m = row['mois']
        mois_labels.append(f"{mois_fr[m.month - 1]} {m.year}")
        mois_data.append(row['total'])

    # ── 2. Demandes par type de travaux ──
    types_qs = (
        Demande.objects
        .values('type_travaux')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    TYPES_MAP = {
        'electricite':  'Électricité',
        'plomberie':    'Plomberie',
        'construction': 'Construction',
        'soudure':      'Soudure',
        'peinture':     'Peinture',
        'menuiserie':   'Menuiserie',
        'surveillance': 'Surveillance',
        'autre':        'Autre',
    }
    types_labels = [TYPES_MAP.get(r['type_travaux'], r['type_travaux']) for r in types_qs]
    types_data   = [r['total'] for r in types_qs]

    # ── 3. Revenus paiements par mois (6 derniers mois) ──
    revenus_qs = (
        Paiement.objects
        .filter(statut='valide', created_at__date__gte=il_y_a_6_mois)
        .annotate(mois=TruncMonth('created_at'))
        .values('mois')
        .annotate(total=Sum('montant'))
        .order_by('mois')
    )
    rev_labels = []
    rev_data   = []
    for row in revenus_qs:
        m = row['mois']
        rev_labels.append(f"{mois_fr[m.month - 1]} {m.year}")
        rev_data.append(float(row['total'] or 0))

    # ── 4. Taux de complétion ──
    total_dem   = Demande.objects.count() or 1
    terminees   = Demande.objects.filter(statut='terminee').count()
    en_cours    = Demande.objects.filter(statut='en_cours').count()
    en_attente  = Demande.objects.filter(statut='en_attente').count()
    annulees    = Demande.objects.filter(statut='annulee').count()
    taux_completion = round((terminees / total_dem) * 100, 1)

    # ── 5. Techniciens les plus actifs (top 5) ──
    top_tech = (
        Demande.objects
        .filter(technicien__isnull=False)
        .values('technicien__prenom', 'technicien__nom', 'technicien__specialite')
        .annotate(nb=Count('id'), terminees=Count('id', filter=__import__('django.db.models', fromlist=['Q']).Q(statut='terminee')))
        .order_by('-nb')[:5]
    )
    tech_labels = [f"{t['technicien__prenom']} {t['technicien__nom']}" for t in top_tech]
    tech_data   = [t['nb'] for t in top_tech]
    tech_terminees = [t['terminees'] for t in top_tech]

    # ── 6. Demandes par localité (top 8) ──
    localites_qs = (
        Demande.objects
        .values('localite')
        .annotate(total=Count('id'))
        .order_by('-total')[:8]
    )
    loc_labels = [r['localite'] for r in localites_qs]
    loc_data   = [r['total'] for r in localites_qs]

    # ── 7. Délai moyen d'intervention (par mois) ──
    # Délai = date_intervention - date_creation (pour les demandes terminées avec date_intervention)
    delai_qs = (
        Demande.objects
        .filter(statut='terminee', date_intervention__isnull=False, date_creation__date__gte=il_y_a_6_mois)
        .annotate(mois=TruncMonth('date_creation'))
        .values('mois')
        .annotate(nb=Count('id'))
        .order_by('mois')
    )
    delai_labels = []
    delai_data   = []
    for row in delai_qs:
        m = row['mois']
        delai_labels.append(f"{mois_fr[m.month - 1]}")
        # Calcul approximatif côté Python
        dems = Demande.objects.filter(
            statut='terminee',
            date_intervention__isnull=False,
            date_creation__month=m.month,
            date_creation__year=m.year,
        )
        delais = [(d.date_intervention - d.date_creation.date()).days for d in dems if d.date_intervention]
        delai_data.append(round(sum(delais) / len(delais), 1) if delais else 0)

    # ── 8. Heatmap : demandes par jour de la semaine × heure ──
    heatmap_data = []
    jours = ['Lun','Mar','Mer','Jeu','Ven','Sam','Dim']
    for jour_idx in range(7):
        for heure in range(0, 24, 2):
            nb = Demande.objects.filter(
                date_creation__week_day=(jour_idx + 2) % 7 + 1,
                date_creation__hour__gte=heure,
                date_creation__hour__lt=heure + 2,
            ).count()
            heatmap_data.append({'x': heure, 'y': jour_idx, 'v': nb})

    # ── KPIs rapides ──
    ce_mois = Demande.objects.filter(
        date_creation__month=today.month,
        date_creation__year=today.year
    ).count()
    mois_precedent = Demande.objects.filter(
        date_creation__month=(today.replace(day=1) - timedelta(days=1)).month,
        date_creation__year=(today.replace(day=1) - timedelta(days=1)).year
    ).count()
    evolution = round(((ce_mois - mois_precedent) / max(mois_precedent, 1)) * 100, 1)

    revenus_mois = Paiement.objects.filter(
        statut='valide',
        created_at__month=today.month,
        created_at__year=today.year
    ).aggregate(t=Sum('montant'))['t'] or 0

    return render(request, 'dashboard/statistiques.html', {
        # Line chart
        'mois_labels':      json.dumps(mois_labels),
        'mois_data':        json.dumps(mois_data),
        # Types
        'types_labels':     json.dumps(types_labels),
        'types_data':       json.dumps(types_data),
        # Revenus
        'rev_labels':       json.dumps(rev_labels),
        'rev_data':         json.dumps(rev_data),
        # Completion
        'taux_completion':  taux_completion,
        'terminees':        terminees,
        'en_cours':         en_cours,
        'en_attente':       en_attente,
        'annulees':         annulees,
        'total_dem':        total_dem,
        # Techniciens
        'tech_labels':      json.dumps(tech_labels),
        'tech_data':        json.dumps(tech_data),
        'tech_terminees':   json.dumps(tech_terminees),
        'top_tech':         list(top_tech),
        # Localités
        'loc_labels':       json.dumps(loc_labels),
        'loc_data':         json.dumps(loc_data),
        # Délai
        'delai_labels':     json.dumps(delai_labels),
        'delai_data':       json.dumps(delai_data),
        # Heatmap
        'heatmap_data':     json.dumps(heatmap_data),
        # KPIs
        'ce_mois':          ce_mois,
        'evolution':        evolution,
        'revenus_mois':     revenus_mois,
    })
    
    
    
# ─────────────────────────────────────────────
#  Modifier un paiement (admin)
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def paiement_modifier(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:paiement_detail', pk=pk)

    paiement = get_object_or_404(Paiement, pk=pk)

    # Champs modifiables
    statut          = request.POST.get('statut', paiement.statut)
    montant         = request.POST.get('montant', paiement.montant)
    mode_paiement   = request.POST.get('mode_paiement', paiement.mode_paiement)
    fournisseur     = request.POST.get('fournisseur_mobile', paiement.fournisseur_mobile or '')
    numero_mobile   = request.POST.get('numero_mobile', paiement.numero_mobile or '')
    note_admin      = request.POST.get('note_admin', paiement.note_admin or '').strip()

    paiement.statut         = statut
    paiement.mode_paiement  = mode_paiement
    paiement.note_admin     = note_admin or None

    try:
        paiement.montant = float(montant)
    except (ValueError, TypeError):
        pass

    if mode_paiement == 'mobile':
        paiement.fournisseur_mobile = fournisseur or None
        paiement.numero_mobile      = numero_mobile or None

    # Si l'admin change le statut manuellement vers validé/rejeté, on enregistre qui l'a fait
    if statut in ('valide', 'rejete') and not paiement.valide_par:
        from django.utils import timezone
        paiement.valide_par      = request.user
        paiement.date_validation = timezone.now()

    paiement.save()
    messages.success(request, f'Paiement #{pk} modifié avec succès.')
    return redirect('dashboard:paiement_detail', pk=pk)


# ─────────────────────────────────────────────
#  Supprimer un paiement (admin)
# ─────────────────────────────────────────────
@login_required
@user_passes_test(is_admin)
def paiement_supprimer(request, pk):
    if request.method != 'POST':
        return redirect('dashboard:paiement_detail', pk=pk)

    paiement = get_object_or_404(Paiement, pk=pk)
    paiement.delete()
    messages.success(request, f'Paiement #{pk} supprimé définitivement.')
    return redirect('dashboard:paiements')





"""
Ajout dans apps/dashboard/views_paiements.py

Vue : notifier_paiement
- L'admin envoie une notification au client pour lui demander de procéder au paiement
- Crée une Notification dans la base avec type 'paiement_requis'
- Redirige vers le détail du paiement avec message de succès
"""

# ─────────────────────────────────────────────
#  Notifier le client — ajouter dans views_paiements.py
# ─────────────────────────────────────────────
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages as django_messages

# Import ton modèle Notification (déjà dans apps.demandes)
from apps.demandes.models import Paiement, InstructionPaiement, Notification

is_admin = lambda u: u.is_staff or u.is_superuser


@login_required
@user_passes_test(is_admin)
def paiement_notifier(request, pk):
    """
    POST /dashboard/paiements/<pk>/notifier/
    Envoie une notification au client pour lui demander de régler son paiement.
    """
    if request.method != 'POST':
        return redirect('dashboard:paiement_detail', pk=pk)

    paiement = get_object_or_404(Paiement, pk=pk)
    message_custom = request.POST.get('message', '').strip()

    # Message par défaut si l'admin ne saisit rien
    instruction = InstructionPaiement.objects.filter(
        mode_paiement=paiement.mode_paiement, actif=True
    ).first()

    mode_label = {
        'carte': 'Carte Bancaire',
        'livraison': 'À la Livraison',
        'mobile': 'Mobile Money',
    }.get(paiement.mode_paiement, paiement.mode_paiement)

    if not message_custom:
        if instruction:
            message_custom = (
                f"Votre paiement pour la demande #{paiement.commande_id} est en attente. "
                f"Mode choisi : {mode_label}. "
                f"{instruction.description}"
            )
        else:
            message_custom = (
                f"Votre paiement pour la demande #{paiement.commande_id} est en attente. "
                f"Veuillez procéder au règlement via {mode_label}."
            )

    # Créer la notification pour le client
    Notification.objects.create(
        utilisateur=paiement.client,
        type='paiement_requis',          # type custom — voir note ci-dessous
        titre='Action requise : Paiement',
        message=message_custom,
        lien=f'/demandes/{paiement.commande_id}/' if paiement.commande_id else None,
        paiement=paiement,               # FK optionnelle — voir note
    )

    django_messages.success(
        request,
        f'Notification envoyée à {paiement.client.get_full_name() or paiement.client.username}.'
    )
    return redirect('dashboard:paiement_detail', pk=pk)
@login_required
@user_passes_test(is_admin)
def client_view_readonly(request, client_pk):

    client        = get_object_or_404(User, pk=client_pk, is_staff=False)
    demandes      = Demande.objects.filter(client=client).order_by('-date_creation')
    paiements     = Paiement.objects.filter(client=client).order_by('-created_at')
    notifications = Notification.objects.filter(utilisateur=client).order_by('-date')

    if request.method == 'POST':
        paiement_id = request.POST.get('paiement_id')
        action      = request.POST.get('action')

        paiement = get_object_or_404(Paiement, pk=paiement_id, client=client)

        if action == 'donner' and paiement.statut == 'valide':
            Paiement.objects.filter(pk=paiement.pk).update(acces_dashboard_vip=True)
            messages.success(
                request,
                f"✅ Accès VIP activé pour {client.get_full_name() or client.username} "
                f"— Paiement #PAY-{paiement.pk:05d}"
            )

        elif action == 'retirer':
            Paiement.objects.filter(pk=paiement.pk).update(acces_dashboard_vip=False)
            messages.warning(
                request,
                f"🔒 Accès VIP retiré — Paiement #PAY-{paiement.pk:05d}"
            )

        else:
            messages.error(
                request,
                "⚠️ Le paiement doit d'abord être validé avant de donner l'accès VIP."
            )

        # ✅ Reste sur le dashboard admin — section paiements
        return redirect(f'/dashboard/clients/{client_pk}/voir/#paiements')

    return render(request, 'dashboard/client_view_readonly.html', {
        'client':             client,
        'demandes':           demandes,
        'paiements':          paiements,
        'notifications':      notifications,
        'nb_demandes':        demandes.count(),
        'nb_paiements':       paiements.count(),
        'nb_notifs_non_lues': notifications.filter(lue=False).count(),
    })
 
# ============================================================
#  3. views.py — Vue client : payment_success (dashboard VIP)
#     Accessible UNIQUEMENT si l'admin a activé acces_dashboard_vip
# ============================================================
 
@login_required
def payment_success(request, paiement_id):
    """
    Dashboard VIP — accessible uniquement si :
    - Le paiement appartient au client connecté
    - Le statut est 'valide'
    - L'admin a coché acces_dashboard_vip = True
    """
    paiement = get_object_or_404(
        Paiement,
        pk=paiement_id,
        client=request.user
    )
 
    # ── Vérification statut ──
    if paiement.statut != 'valide':
        messages.warning(
            request,
            "⏳ Votre paiement est en cours de validation. "
            "Vous aurez accès au dashboard dès confirmation."
        )
        return redirect('demandes:mes_demandes')
 
    # ── Vérification accès VIP donné par l'admin ──
    if not paiement.acces_dashboard_vip:
        messages.info(
            request,
            "🔒 L'accès à ce dashboard sera activé par notre équipe très prochainement."
        )
        return redirect('demandes:mes_demandes')
 
    return render(request, 'paiements/payment_success.html', {
        'paiement': paiement,
    })
 
 # ============================================================
#  4. Vue liste des paiements VIP du client
# ============================================================
 
@login_required
def mes_paiements_valides(request):
    """
    Affiche uniquement les paiements pour lesquels
    l'admin a accordé l'accès VIP.
    """
    paiements = Paiement.objects.filter(
        client=request.user,
        statut='valide',
        acces_dashboard_vip=True          # ← seulement ceux autorisés par l'admin
    ).select_related('commande').order_by('-date_validation')
 
    return render(request, 'paiements/mes_paiements.html', {
        'paiements': paiements,
    })

from apps.core.models import ContactMessage
 
@login_required
@user_passes_test(is_admin)
def contact_messages(request):
    msgs = ContactMessage.objects.all()
    return render(request, 'dashboard/contact_messages.html', {
        'contact_messages':  msgs,
        'total':             msgs.count(),
        'messages_non_lus':  msgs.filter(lu=False).count(),
    })
 
@login_required
@user_passes_test(is_admin)
def contact_marquer_lu(request, pk):
    from django.shortcuts import get_object_or_404
    msg = get_object_or_404(ContactMessage, pk=pk)
    msg.lu = True
    msg.save()
    return redirect('dashboard:contact_messages')
 
@login_required
@user_passes_test(is_admin)
def contact_supprimer(request, pk):
    from django.shortcuts import get_object_or_404
    msg = get_object_or_404(ContactMessage, pk=pk)
    if request.method == 'POST':
        msg.delete()
    return redirect('dashboard:contact_messages')



# ── À AJOUTER à la fin de apps/dashboard/views_paiements.py ──
# pip install openpyxl  (déjà inclus dans Django si pas installé)

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def hfill(color): return PatternFill("solid", fgColor=color)
def thin():
    s = Side(style='thin', color="E2E8F0")
    return Border(top=s, bottom=s, left=s, right=s)
def bold(size=10, color="000000"): return Font(bold=True, size=size, color=color, name="Arial")
def reg(size=10, color="000000"):  return Font(bold=False, size=size, color=color, name="Arial")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def right():  return Alignment(horizontal="right",  vertical="center")


@login_required
@user_passes_test(is_admin)
def export_paiements_excel(request):
    """
    GET /dashboard/paiements/export/
    Génère et télécharge un fichier Excel complet de l'historique des paiements.
    """
    # ── Récupérer les paiements ──
    paiements = Paiement.objects.select_related('client', 'commande').order_by('-created_at')

    wb = Workbook()

    # ═══════════════════════════════════
    #  FEUILLE 1 — HISTORIQUE
    # ═══════════════════════════════════
    ws = wb.active
    ws.title = "Historique Paiements"
    ws.sheet_view.showGridLines = False

    # Bandeau titre
    ws.merge_cells("A1:L2")
    c = ws["A1"]
    c.value = "INFINITY HOME REWIRE & CONSTRUCTION — Historique des Paiements"
    c.font  = bold(14, "FFFFFF")
    c.fill  = hfill("1E3A8A")
    c.alignment = center()
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 12

    # Sous-titre
    ws.merge_cells("A3:L3")
    c = ws["A3"]
    c.value = f"Exporté le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}     |     Total : {paiements.count()} paiement(s)"
    c.font  = reg(9, "64748B")
    c.fill  = hfill("DBEAFE")
    c.alignment = center()
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 8

    # En-têtes
    headers    = ["N°", "Date", "Client", "Email", "Demande N°",
                  "Type travaux", "Mode paiement", "Montant (BIF)",
                  "Statut", "Validé par", "Date validation", "Note admin"]
    col_widths = [7, 14, 24, 28, 13, 18, 18, 16, 13, 18, 16, 32]

    ws.row_dimensions[5].height = 20
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        col = get_column_letter(i)
        cell = ws[f"{col}5"]
        cell.value = h
        cell.font  = bold(10, "FFFFFF")
        cell.fill  = hfill("2563EB")
        cell.alignment = center()
        cell.border = thin()
        ws.column_dimensions[col].width = w

    # Données réelles
    MODE_LABELS = {
        'carte':     'Carte Bancaire',
        'livraison': 'À la livraison',
        'mobile':    'Mobile Money',
    }
    TYPE_LABELS = {
        'electricite':  'Électricité',
        'plomberie':    'Plomberie',
        'construction': 'Construction',
        'soudure':      'Soudure',
        'peinture':     'Peinture',
        'menuiserie':   'Menuiserie',
        'surveillance': 'Surveillance',
        'autre':        'Autre',
    }

    for row_idx, p in enumerate(paiements, start=6):
        ws.row_dimensions[row_idx].height = 18
        bg = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"

        statut_label = {
            'valide':     'Validé',
            'en_attente': 'En attente',
            'rejete':     'Rejeté',
        }.get(p.statut, p.statut)

        row = [
            row_idx - 5,
            p.created_at.strftime('%d/%m/%Y') if p.created_at else '',
            p.client.get_full_name() or p.client.username,
            p.client.email or '',
            f"DEM-{p.commande_id}" if p.commande_id else '',
            TYPE_LABELS.get(p.commande.type_travaux, '') if p.commande else '',
            MODE_LABELS.get(p.mode_paiement, p.mode_paiement),
            float(p.montant) if p.montant else 0,
            statut_label,
            p.valide_par.get_full_name() if p.valide_par else '',
            p.date_validation.strftime('%d/%m/%Y') if hasattr(p, 'date_validation') and p.date_validation else '',
            p.note_admin or '',
        ]

        for col_idx, val in enumerate(row, 1):
            col  = get_column_letter(col_idx)
            cell = ws[f"{col}{row_idx}"]
            cell.value  = val
            cell.border = thin()

            if col_idx == 9:  # Statut
                if p.statut == 'valide':
                    cell.font = bold(9, "065F46"); cell.fill = hfill("D1FAE5")
                elif p.statut == 'rejete':
                    cell.font = bold(9, "991B1B"); cell.fill = hfill("FEE2E2")
                else:
                    cell.font = bold(9, "92400E"); cell.fill = hfill("FEF3C7")
                cell.alignment = center()
            elif col_idx == 8:
                cell.font = bold(10, "1E3A8A")
                cell.fill = hfill(bg)
                cell.alignment = right()
                cell.number_format = '#,##0'
            elif col_idx == 1:
                cell.font = bold(9, "2563EB")
                cell.fill = hfill(bg)
                cell.alignment = center()
            else:
                cell.font = reg(9)
                cell.fill = hfill(bg)
                cell.alignment = left()

    last_row = 5 + paiements.count()

    # Ligne TOTAL
    total_row = last_row + 1
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(f"A{total_row}:G{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "TOTAL PAIEMENTS VALIDÉS"
    c.font  = bold(10, "FFFFFF")
    c.fill  = hfill("1E3A8A")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = thin()

    c2 = ws[f"H{total_row}"]
    c2.value = f'=SUMIF(I6:I{last_row},"Validé",H6:H{last_row})'
    c2.font  = bold(11, "FFFFFF")
    c2.fill  = hfill("1E3A8A")
    c2.alignment = right()
    c2.number_format = '#,##0'
    c2.border = thin()

    for col in range(9, 13):
        cell = ws[f"{get_column_letter(col)}{total_row}"]
        cell.fill = hfill("1E3A8A")
        cell.border = thin()

    ws.freeze_panes = "A6"

    # ═══════════════════════════════════
    #  FEUILLE 2 — RÉSUMÉ
    # ═══════════════════════════════════
    from django.db.models import Sum as DSum, Count
    ws2 = wb.create_sheet("Résumé")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D2")
    c = ws2["A1"]
    c.value = "RÉSUMÉ DES PAIEMENTS — Infinity Home"
    c.font  = bold(13, "FFFFFF")
    c.fill  = hfill("1E3A8A")
    c.alignment = center()
    ws2.row_dimensions[1].height = 22

    total_valide    = paiements.filter(statut='valide').aggregate(t=DSum('montant'))['t'] or 0
    total_attente   = paiements.filter(statut='en_attente').count()
    total_rejete    = paiements.filter(statut='rejete').count()
    total_val_count = paiements.filter(statut='valide').count()

    summary = [
        ("Total paiements",              paiements.count(),  "DBEAFE", "1E3A8A"),
        ("Paiements validés",            total_val_count,    "D1FAE5", "065F46"),
        ("Paiements en attente",         total_attente,      "FEF3C7", "92400E"),
        ("Paiements rejetés",            total_rejete,       "FEE2E2", "991B1B"),
        ("Montant total validé (BIF)",   f"{total_valide:,.0f}", "DBEAFE", "1E3A8A"),
    ]

    for i, (label, val, bg, tc) in enumerate(summary, start=4):
        ws2.row_dimensions[i].height = 22
        ws2.merge_cells(f"A{i}:C{i}")
        c1 = ws2[f"A{i}"]
        c1.value = label
        c1.font  = bold(10, "1E293B")
        c1.fill  = hfill(bg)
        c1.alignment = left()
        c1.border = thin()

        ws2.merge_cells(f"D{i}:E{i}")
        c2 = ws2[f"D{i}"]
        c2.value = val
        c2.font  = bold(12, tc)
        c2.fill  = hfill(bg)
        c2.alignment = right()
        c2.border = thin()

    for col, w in zip(["A","B","C","D","E"], [26, 16, 16, 18, 16]):
        ws2.column_dimensions[col].width = w

    # ── Réponse HTTP ──
    filename = f"Historique_Paiements_InfinityHome_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



# ── À AJOUTER à la fin de apps/dashboard/views_paiements.py ──

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (BaseDocTemplate, Frame, PageTemplate,
                                 Paragraph, Spacer, Table, TableStyle,
                                 HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import io

# ── Couleurs ──
NAVY   = colors.HexColor("#1E3A8A")
BLUE   = colors.HexColor("#2563EB")
LBLUE  = colors.HexColor("#DBEAFE")
LGRAY  = colors.HexColor("#F8FAFC")
GRAY   = colors.HexColor("#E2E8F0")
DGRAY  = colors.HexColor("#64748B")
TEXT   = colors.HexColor("#0F172A")
GREEN  = colors.HexColor("#D1FAE5")
DGREEN = colors.HexColor("#065F46")
ORANGE = colors.HexColor("#FEF3C7")
DORANGE= colors.HexColor("#92400E")
RED    = colors.HexColor("#FEE2E2")
DRED   = colors.HexColor("#991B1B")
WHITE  = colors.white
W, H   = A4


def _draw_header_footer(c, doc):
    """Dessine l'en-tête et le pied de page sur chaque page."""
    c.saveState()

    # ── HEADER ──
    c.setFillColor(NAVY)
    c.rect(0, H - 50*mm, W, 50*mm, fill=1, stroke=0)

    # Cercle logo IH
    c.setFillColor(BLUE)
    c.circle(25*mm, H - 25*mm, 14*mm, fill=1, stroke=0)
    c.setStrokeColor(WHITE)
    c.setLineWidth(1)
    c.circle(25*mm, H - 25*mm, 11*mm, fill=0, stroke=1)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(25*mm, H - 28*mm, "IH")

    # Nom entreprise
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(44*mm, H - 20*mm, "INFINITY HOME")
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#93C5FD"))
    c.drawString(44*mm, H - 28*mm, "REWIRE & CONSTRUCTION")

    # Séparateur
    c.setStrokeColor(colors.HexColor("#3B82F6"))
    c.setLineWidth(0.5)
    c.line(44*mm, H - 32*mm, W - 15*mm, H - 32*mm)

    # Contacts
    c.setFont("Helvetica", 7.5)
    c.setFillColor(colors.HexColor("#BFDBFE"))
    c.drawString(44*mm, H - 38*mm, "+257 66 033 033")
    c.drawString(82*mm, H - 38*mm, "|  info@infinityhome.bi")
    c.drawString(132*mm, H - 38*mm, "|  Bujumbura, Burundi")

    # Titre + date (droite)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(W - 15*mm, H - 22*mm, "HISTORIQUE DES PAIEMENTS")
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#93C5FD"))
    c.drawRightString(W - 15*mm, H - 30*mm,
                      f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")

    # ── FOOTER ──
    c.setFillColor(NAVY)
    c.rect(0, 0, W, 14*mm, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica", 7.5)
    c.drawString(15*mm, 8*mm,
                 "© 2025 Infinity Home Rewire & Construction — Tous droits réservés")
    c.drawString(15*mm, 4*mm,
                 "Bujumbura, Burundi  |  +257 66 033 033  |  info@infinityhome.bi")
    c.setFont("Helvetica-Bold", 8)
    c.drawRightString(W - 15*mm, 6*mm, f"Page {doc.page}")

    c.restoreState()


def _ps(name, size=9, color=TEXT, bold=False, align=TA_LEFT, leading=None):
    return ParagraphStyle(
        name,
        fontName="Helvetica-Bold" if bold else "Helvetica",
        fontSize=size,
        textColor=color,
        alignment=align,
        leading=leading or size * 1.4,
    )


@login_required
@user_passes_test(is_admin)
def export_paiements_pdf(request):
    """
    GET /dashboard/paiements/pdf/
    Génère et télécharge un PDF de l'historique des paiements.
    """
    paiements_qs = Paiement.objects.select_related(
        'client', 'commande', 'valide_par'
    ).order_by('-created_at')

    TYPE_LABELS = {
        'electricite': 'Électricité', 'plomberie': 'Plomberie',
        'construction': 'Construction', 'soudure': 'Soudure',
        'peinture': 'Peinture', 'menuiserie': 'Menuiserie',
        'surveillance': 'Surveillance', 'autre': 'Autre',
    }
    MODE_LABELS = {
        'carte': 'Carte Bancaire',
        'livraison': 'À la livraison',
        'mobile': 'Mobile Money',
    }
    STATUT_LABELS = {
        'valide': 'Validé',
        'en_attente': 'En attente',
        'rejete': 'Rejeté',
    }

    # KPIs
    total         = paiements_qs.count()
    nb_valides    = paiements_qs.filter(statut='valide').count()
    nb_attente    = paiements_qs.filter(statut='en_attente').count()
    nb_rejetes    = paiements_qs.filter(statut='rejete').count()
    from django.db.models import Sum as DSum
    total_montant = paiements_qs.filter(statut='valide').aggregate(
        t=DSum('montant'))['t'] or 0

    # ── Buffer mémoire ──
    buffer = io.BytesIO()

    doc = BaseDocTemplate(
        buffer, pagesize=A4,
        topMargin=55*mm, bottomMargin=20*mm,
        leftMargin=15*mm, rightMargin=15*mm,
        title="Historique des Paiements — Infinity Home",
        author="Infinity Home Rewire & Construction",
        subject="Rapport des paiements",
        creator="Plateforme Infinity Home",
    )
    frame = Frame(
        doc.leftMargin, doc.bottomMargin,
        W - doc.leftMargin - doc.rightMargin,
        H - doc.topMargin - doc.bottomMargin,
        id='main'
    )
    doc.addPageTemplates([
        PageTemplate(id='main', frames=[frame], onPage=_draw_header_footer)
    ])

    story = []

    # Titre
    story.append(Paragraph("Rapport des Paiements",
                            _ps("t1", 14, NAVY, True)))
    story.append(Spacer(1, 2*mm))
    story.append(HRFlowable(width="100%", thickness=2,
                             color=BLUE, spaceAfter=4*mm))

    # ── KPIs ──
    kpi_data = [[
        Paragraph(f'<b>{total}</b><br/><font size="7" color="#64748B">Total</font>',
                  _ps("k", 18, NAVY, True, TA_CENTER, 22)),
        Paragraph(f'<b>{nb_valides}</b><br/><font size="7" color="#065F46">Validés</font>',
                  _ps("k2", 18, DGREEN, True, TA_CENTER, 22)),
        Paragraph(f'<b>{nb_attente}</b><br/><font size="7" color="#92400E">En attente</font>',
                  _ps("k3", 18, DORANGE, True, TA_CENTER, 22)),
        Paragraph(f'<b>{nb_rejetes}</b><br/><font size="7" color="#991B1B">Rejetés</font>',
                  _ps("k4", 18, DRED, True, TA_CENTER, 22)),
        Paragraph(f'<b>{int(total_montant):,}</b><br/><font size="7" color="#64748B">BIF validés</font>',
                  _ps("k5", 13, BLUE, True, TA_CENTER, 18)),
    ]]
    kpi_t = Table(kpi_data, colWidths=[32*mm, 32*mm, 32*mm, 32*mm, 42*mm])
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(0,0), LBLUE),
        ("BACKGROUND",   (1,0),(1,0), GREEN),
        ("BACKGROUND",   (2,0),(2,0), ORANGE),
        ("BACKGROUND",   (3,0),(3,0), RED),
        ("BACKGROUND",   (4,0),(4,0), LBLUE),
        ("BOX",          (0,0),(-1,-1), 0.5, GRAY),
        ("INNERGRID",    (0,0),(-1,-1), 0.5, GRAY),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story.append(kpi_t)
    story.append(Spacer(1, 6*mm))

    # ── Tableau paiements ──
    story.append(Paragraph("Détail des paiements",
                            _ps("t2", 11, NAVY, True)))
    story.append(Spacer(1, 2*mm))

    # Colonne "Titre demande" plus large
    col_w = [10*mm, 18*mm, 30*mm, 38*mm, 20*mm, 20*mm, 22*mm, 14*mm]

    def hdr(txt):
        return Paragraph(f"<b>{txt}</b>",
                         _ps("h", 7.5, WHITE, True, TA_CENTER))

    rows = [[
        hdr("N°"), hdr("Date"), hdr("Client"), hdr("Titre demande"),
        hdr("Type de Construction"), hdr("Mode de paiement"), hdr("Montant"), hdr("Statut")
    ]]

    statut_bg_map = []
    for i, p in enumerate(paiements_qs, start=1): 
        st      = p.statut
        st_lbl  = STATUT_LABELS.get(st, st)
        montant_val = float(p.montant) if p.montant else 0

        if st == 'valide':   sc, sbg = DGREEN,  GREEN
        elif st == 'rejete': sc, sbg = DRED,    RED
        else:                sc, sbg = DORANGE, ORANGE
        statut_bg_map.append((i, sbg))

        # ── Titre de la demande ──
        titre_demande = ''
        if p.commande:
            titre_demande = p.commande.titre or f"Demande #{p.commande_id}"

        type_txt = TYPE_LABELS.get(
            p.commande.type_travaux, '') if p.commande else ''
        mode_txt = MODE_LABELS.get(p.mode_paiement, p.mode_paiement)

        rows.append([
            Paragraph(str(i),
                      _ps("d", 8, BLUE, True, TA_CENTER)),
            Paragraph(p.created_at.strftime('%d/%m/%Y') if p.created_at else '',
                      _ps("d", 8, DGRAY, False, TA_CENTER)),
            Paragraph(p.client.get_full_name() or p.client.username,
                      _ps("d", 8, TEXT, False, TA_LEFT)),
            Paragraph(titre_demande,
                      _ps("d", 8, TEXT, False, TA_LEFT)),
            Paragraph(type_txt,
                      _ps("d", 8, TEXT, False, TA_LEFT)),
            Paragraph(mode_txt,
                      _ps("d", 8, DGRAY, False, TA_LEFT)),
            Paragraph(f"{montant_val:,.0f} BIF",
                      _ps("d", 8, NAVY, True, TA_RIGHT)),
            Paragraph(f"<b>{st_lbl}</b>",
                      _ps("d", 7.5, sc, True, TA_CENTER)),
        ])

    # Ligne total
    rows.append([
        Paragraph("", _ps("x")),
        Paragraph("", _ps("x")),
        Paragraph("", _ps("x")),
        Paragraph("", _ps("x")),
        Paragraph("", _ps("x")),
        Paragraph("<b>TOTAL VALIDÉ</b>",
                  _ps("tot", 8.5, WHITE, True, TA_RIGHT)),
        Paragraph(f"<b>{int(total_montant):,} BIF</b>",
                  _ps("tot", 8.5, WHITE, True, TA_RIGHT)),
        Paragraph("", _ps("x")),
    ])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    ts  = [
        ("BACKGROUND",     (0,0),  (-1,0),  NAVY),
        ("ROWBACKGROUNDS", (0,1),  (-1,-2), [WHITE, LGRAY]),
        ("BACKGROUND",     (0,-1), (-1,-1), BLUE),
        ("LINEBELOW",      (0,0),  (-1,0),  1,   BLUE),
        ("LINEBELOW",      (0,1),  (-1,-2), 0.3, GRAY),
        ("LINEABOVE",      (0,-1), (-1,-1), 1.5, NAVY),
        ("BOX",            (0,0),  (-1,-1), 0.5, GRAY),
        ("TOPPADDING",     (0,0),  (-1,-1), 5),
        ("BOTTOMPADDING",  (0,0),  (-1,-1), 5),
        ("LEFTPADDING",    (0,0),  (-1,-1), 4),
        ("RIGHTPADDING",   (0,0),  (-1,-1), 4),
        ("VALIGN",         (0,0),  (-1,-1), "MIDDLE"),
    ]
    for row_i, bg in statut_bg_map:
        ts.append(("BACKGROUND", (7, row_i), (7, row_i), bg))
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)

    # Note pied
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=GRAY, spaceAfter=2*mm))
    story.append(Paragraph(
        "Document généré automatiquement par la plateforme Infinity Home. "
        "Pour toute question, contactez l'administration.",
        _ps("note", 7.5, DGRAY, False, TA_CENTER)
    ))

    doc.build(story)
    buffer.seek(0)  # ← CORRIGÉ (était 1)

    filename = f"InfinityHome_Paiements_{datetime.now().strftime('%d-%m-%Y')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response










from django.contrib.auth.models import User, Group
from apps.accounts.models import Profil
from apps.techniciens.models import Technicien, SPECIALITES
from itertools import groupby
 
 
@admin_required
def clients_par_localite(request):
    """
    Affiche les clients regroupés par localité.
    Utilise à la fois les groupes Django ET le champ localite du Profil.
    """
    # Récupère toutes les localités distinctes
    localites = (
        Profil.objects
        .exclude(localite='')
        .values_list('localite', flat=True)
        .distinct()
        .order_by('localite')
    )
 
    groupes = []
    for loc in localites:
        clients = User.objects.filter(
            is_staff=False,
            profil__localite=loc
        ).select_related('profil').order_by('last_name', 'first_name')
 
        if clients.exists():
            groupes.append({
                'localite': loc,
                'clients': clients,
                'count': clients.count(),
                # Groupe Django correspondant
                'groupe_django': Group.objects.filter(
                    name=f"Localité : {loc}"
                ).first(),
            })
 
    # Clients sans localité
    sans_localite = User.objects.filter(
        is_staff=False
    ).filter(
        profil__localite=''
    ).select_related('profil')
 
    return render(request, 'dashboard/clients_par_localite.html', {
        'groupes':        groupes,
        'sans_localite':  sans_localite,
        'total_clients':  User.objects.filter(is_staff=False).count(),
        'total_localites': len(groupes),
    })
 
 
@admin_required
def techniciens_par_specialite(request):
    """
    Affiche les techniciens regroupés par spécialité.
    """
    SPEC_DICT = dict(SPECIALITES)
    groupes = []
 
    for code, label in SPECIALITES:
        techs = Technicien.objects.filter(
            specialite=code, actif=True
        ).order_by('nom')
 
        # Inclure même les spécialités sans techniciens actifs
        groupes.append({
            'code':           code,
            'label':          label,
            'techniciens':    techs,
            'count':          techs.count(),
            'disponibles':    techs.filter(disponibilite='disponible').count(),
            'groupe_django':  Group.objects.filter(
                name=f"Spécialité : {label}"
            ).first(),
        })
 
    # Trier par nombre de techniciens décroissant
    groupes.sort(key=lambda x: x['count'], reverse=True)
 
    return render(request, 'dashboard/techniciens_par_specialite.html', {
        'groupes':           groupes,
        'total_techniciens': Technicien.objects.filter(actif=True).count(),
        'total_specialites': len([g for g in groupes if g['count'] > 0]),
    })
 